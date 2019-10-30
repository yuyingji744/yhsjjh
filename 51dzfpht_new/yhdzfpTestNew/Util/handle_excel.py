import os
import sys
import openpyxl
base_path = os.getcwd()
sys.path.append(base_path)

# 定义操作Excel表格的类
class HandleExcel:
    def load_excel(self):
        '''定义加载excel的方法'''
        open_excel = openpyxl.load_workbook('E:/yhdzfpTestNew/Config/yhdzfp.xlsx')
        return open_excel

    def get_sheet_data(self, index=None):
        '''获取excel表格里所有sheet的内容,传入参数index'''
        sheet_name = self.load_excel().sheetnames
        # 如果index不赋值，默认获取sheet1的内容
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, column):
        '''获取excel表格中某个单元格的数据,传入参数row,column'''
        data = self.get_sheet_data().cell(row=row, column=column).value
        return data

    def get_rows_count(self):
        '''获取excel表格中所有行数'''
        rows_count = self.get_sheet_data().max_row
        return rows_count

    def get_rows_value(self, row):
        '''获取excel表格中某一行的数据，传入参数row'''
        rows_list = []
        for i in self.get_sheet_data()[row]:
            rows_list.append(i.value)
        return rows_list

    def get_rows_number(self, case_id):
        '''获取excel表格中某一行的行号，传入参数case_id'''
        num = 1
        columns_data = self.get_columns_value()
        for cols_data in columns_data:
            if case_id == cols_data:
                return num
            num = num + 1
        return num

    def get_columns_value(self, column=None):
        '''获取excel表格中某一列的数据，传入参数column'''
        columns_list = []
        if column == None:
            column = 'A'
        column_list_data = self.get_sheet_data()[column]
        for i in column_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_excel_all_data(self):
        '''获取excel表格中所有行的数据'''
        data_list = []
        for i in range(self.get_rows_count()):
            data_list.append(self.get_rows_value(i+2))
        return data_list

    def excel_write_data(self, row, column, value):
        '''回写数据到excel表格中，传入参数row,column,value'''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, column, value)
        wb.save('E:/yhdzfpTestNew/Config/yhdzfp.xlsx')

get_excel = HandleExcel()

# if __name__ == '__main__':
    # return_data = get_excel.excel_write_data(2, 10, '通过')
    # row_list_value = get_excel.get_rows_value(2)
#     column_list_value = get_excel.get_columns_value()
#     get_rows = get_excel.get_rows_number('Case_id')
#     all_list_data = get_excel.get_excel_all_data()
    # print(column_list_value)
    # print(column_list_value)
    # print(get_rows)
    # print(row_list_value)
    # print("-------------------------------------------------->")
    # print(all_list_data[0])




